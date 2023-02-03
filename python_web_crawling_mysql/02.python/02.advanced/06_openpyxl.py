# from openpyxl import Workbook
import openpyxl
import pprint


# wb = Workbook()

'''
1. openpy1
'''
# ws = wb.active

# ws.title = 'openpy'

# wb.save('openpy01.xlsx')

# wb.close()


'''
2. openpy2
'''
# ws = wb.create_sheet()

# ws.title = 'openpy2'

# ws3 = wb.create_sheet('openpy3')

# wb['openpy2'].title = 'new openpy2' # 접근하고자 하는 시트의 이름을 키로 넣어 접근한다.

# ws2 = wb['new openpy2'] # 특정 시트를 객체로서 ws2에 할당한다.

# ws2['A3'] = 'eyesone'

# wb.save('openpy01.xlsx')

'''
3. openpy3
'''


# # wb = openpyxl.load_workbook(filename='openpy01.xlsx')
# wb = openpyxl.load_workbook('openpy01.xlsx') # 기존에 있던 파일 가져오기
# ws1 = wb.active # 내가 어떤 시트를 사용 중일지 지정

# # 셀 데이터 입력
# ws1['A1'] = '휫자'
# ws1['a2'] = '치킨'
# ws1['B3'] = '햄버거'
# ws1['B5'] = '마라탕'

# wb.save('openpy01.xlsx')

# # 셀 데이터 출력
# print(ws1['A1'].value)
# print(ws1['C1'].value) # 없는 값은 None Type 출력

# # 셀에 숫자로 접근하기
# # ws.cell(r, c)

# print(ws1.cell(3, 2).value)

# # 구구단
# ws2 = wb.create_sheet('gugu')

# for r in range(1, 10):
#     for c in range(1, 10):
#         # ws2.cell(r, c, r * c)
#         ws2.cell(r, c).value = r * c

# wb.save('openpy01.xlsx')


# openpy 4

wb = openpyxl.load_workbook(filename='./resource/sample.xlsx')

ws = wb.active

max_row = ws.max_row # 최대 행 위치
max_col = ws.max_column # 최대 열 위치
min_row = ws.min_row # 최소 행 위치
min_col = ws.min_column # 최소 열 위치


# for i in range(1, max_row + 1):
#     for j in range(1, max_col + 1):
#         print(ws.cell(i, j).value)

'''
iter_range
row_range = ws[1:4]
col_range = ws['A:B']
'''

# e.g. 연(A열), 월(B열)
# col_range = ws['A:B'] # A ~ B열에 대한 정보를 가져온다. => 한 열씩 자른다.


# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# row_range = ws[1:4] # 한 행씩 리스트로 담는다.

# for rows in row_range:
#     for cell in rows:
#         print(cell.value)

'''
ws.iter_rows, ws.iter_cols
iter_rows => row들을 가져온다, row들을 반복한다.

iter_cols => column들을 가져온다, column들을 반복한다.
'''
# for row in ws.iter_rows(min_row=1, max_row=4):
#     print(row[0].value, row[1].value, row[2].value)

for row in ws.iter_rows(min_row=2):
    if int(row[2].value) >= 3500000:
        print(row[0].value, row[1].value)
